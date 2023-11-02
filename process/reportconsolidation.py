import openpyxl as xl
from utils.file_handling import is_path_exists
from os import path
from process import date_formats
import numpy as np
from operator import itemgetter


class ReportConsolidation():

    def __init__(self, process):
        self.report_data = np.array([])
        (
            self.reports_path,
            self.dunning_path,
            self.dunning_filename,
            self.areas, _
        ) = itemgetter(*process.keys())(process)

    def _checkPathError(self, path) -> bool:
        if not path.exists(path):
            return True
        return False

    def _check_file_error(self, file_with_path) -> bool:
        if not path.exists(file_with_path):
            return True
        return False

    def _read_generated_files(self, generated_file_with_path, filename):
        if not self._check_file_error(generated_file_with_path):
            wb = xl.load_workbook(generated_file_with_path, enumerate)
            ws = wb.worksheets[0]

            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                self.report_data = [*self.report_data, [c.value for c in r]]

            print(f'{filename} ... DONE!')

    def _consolidate_dunning(self, file_path, filename):

        dunning_file = file_path + filename
        print(f"\nConsolidating to {filename} ...")

        dunning = xl.load_workbook(filename=dunning_file)
        ws = dunning['Consolidated Data']

        pivot_sheet = dunning['Pivot']
        pivot = pivot_sheet._pivots[0]
        pivot.cache.refreshOnLoad = True

        for row_index, row in enumerate(
            ws.iter_rows(min_row=2,
                         max_row=len(self.report_data) + 1, max_col=ws.max_column
                         )
        ):
            col = 0
            for cell in row:
                cell.value = self.report_data[row_index][col]
                col += 1

        dunning.save(dunning_file)
        print(f"\nDONE! Opening {filename} ...")

        # open excel file upon completion
        self._open_excel_file(file_path, filename)

    def _open_excel_file(self, path, filename):
        from utils.file_handling import change_working_dir, run_system_command as open_file
        change_working_dir(path)
        open_file(f'start excel.exe {filename}')

    def execute_process(self):

        if not is_path_exists(self.reports_path):
            print(
                f"\n-Discon_Generated path does not exists: {self.reports_path}]")
            return None

        if not is_path_exists(self.dunning_path):
            print(f"\nReading dunning files from path `{self.dunning_path}`: ")

            for area in self.areas:
                generated_filename = f"""DISCON_LIST_{area}_{date_formats.current_year}
                    {date_formats.month_number}{date_formats.current_day_number}.xlsx"""

                file_with_path = self.dunning_path + generated_filename
                self._read_generated_files(file_with_path, generated_filename)

            self._consolidate_dunning(
                self.dunning_path, self.dunning_filename)

            self.report_data = np.array([])
